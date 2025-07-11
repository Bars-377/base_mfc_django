from celery import shared_task
from asgiref.sync import async_to_sync
import os

import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side
from channels.layers import get_channel_layer

@shared_task
def generate_excel(sid, data):
    """
    Генерация Excel-файла с данными.
    """
    try:
        from data_app.models import Services, Services_Two, Services_Three

        print(f"Начало экспорта для сессии {sid} с данными: {data}")

        contract_date = data[0]

        end_date = data[1]

        query = Services.objects.all()

        # Регулярное выражение для формата "DD.MM.YYYY"
        pattern_dd_mm_yyyy = r'\b\d{2}\.\d{2}\.\d{4}\b'

        # Регулярное выражение для формата "YYYY-MM-DD"
        pattern_yyyy_mm_dd = r'\b\d{4}-\d{2}-\d{2}\b'

        if contract_date == 'No':
            contract_date = None
        if end_date == 'No':
            end_date = None

        def apply_filters(query, filters):
            for filter_func in filters:
                query = filter_func(query)
            return query

        from django.db.models import Q

        filters = []
        if contract_date == 'None' and end_date == 'None':
            filters.append(lambda q: q.exclude(Q(contract_date__regex=pattern_dd_mm_yyyy) | Q(contract_date__regex=pattern_yyyy_mm_dd) | Q(end_date__regex=pattern_dd_mm_yyyy) | Q(end_date__regex=pattern_yyyy_mm_dd)))
        elif contract_date == 'None':
            filters.append(lambda q: q.exclude(Q(contract_date__regex=pattern_dd_mm_yyyy) | Q(contract_date__regex=pattern_yyyy_mm_dd)))
            if end_date:
                filters.append(lambda q: q.filter(end_date__icontains=end_date))
        elif end_date == 'None':
            filters.append(lambda q: q.exclude(Q(end_date__regex=pattern_dd_mm_yyyy) | Q(end_date__regex=pattern_yyyy_mm_dd)))
            if contract_date:
                filters.append(lambda q: q.filter(contract_date__icontains=contract_date))
        elif contract_date and end_date:
            filters.append(lambda q: q.filter(Q(contract_date__icontains=contract_date) | Q(end_date__icontains=end_date)))
        elif contract_date:
            filters.append(lambda q: q.filter(contract_date__icontains=contract_date))
        elif end_date:
            filters.append(lambda q: q.filter(end_date__icontains=end_date))

        query = apply_filters(query, filters)

        # Сортировка
        from django.db.models import IntegerField
        from django.db.models.functions import Cast

        query = query.annotate(
            id_id_int=Cast('id_id', IntegerField())
        ).order_by('id_id_int')

        services = query
        print(f"Получено {len(services)} услуг.")

        df = pd.DataFrame([{
            '№ п/п': service.id_id,
            'Наименование закупки': service.name,
            'Статус': service.status,
            'Способ закупки': service.way,
            'Инициатор закупки (ИТ/АХО)': service.initiator,
            'КЦСР': service.KTSSR,
            'КОСГУ': service.KOSGU,
            'ДопФК': service.DopFC,
            'НМЦК': service.NMCC,
            'Экономия': service.saving,
            'Контрагент': service.counterparty,
            'Реестровый номер контракта (ЕИС)': service.registration_number,
            'Номер контракта': service.contract_number,
            'Дата контракта': service.contract_date,
            'Окончание даты исполнения': service.end_date,
            'Цена контракта': service.contract_price,
            'Исполнение контракта (план) (формула)': service.execution_contract_plan,
            'Остаток контракта с предыдущего года': service.remainder_old_year,
            'Январь (план)': service.january_one,
            'Февраль (план)': service.february,
            'Март (план)': service.march,
            'Апрель (план)': service.april,
            'Май (план)': service.may,
            'Июнь (план)': service.june,
            'Июль (план)': service.july,
            'Август (план)': service.august,
            'Сентябрь (план)': service.september,
            'Октябрь (план)': service.october,
            'Ноябрь (план)': service.november,
            'Декабрь (план)': service.december,
            'Остаток контракта на следующий год': service.january_two,
            'Исполнение контракта (факт) (формула)': service.execution_contract_fact,
            'Оплачено в предыдущем году': service.paid_last_year,
            'Дата оплаты Январь (факт)': service.date_january_one,
            'Сумма оплаты Январь (факт)': service.sum_january_one,
            'Дата оплаты Февраль (факт)': service.date_february,
            'Сумма оплаты Февраль (факт)': service.sum_february,
            'Дата оплаты Март (факт)': service.date_march,
            'Сумма оплаты Март (факт)': service.sum_march,
            'Дата оплаты Апрель (факт)': service.date_april,
            'Сумма оплаты Апрель (факт)': service.sum_april,
            'Дата оплаты Май (факт)': service.date_may,
            'Сумма оплаты Май (факт)': service.sum_may,
            'Дата оплаты Июнь (факт)': service.date_june,
            'Сумма оплаты Июнь (факт)': service.sum_june,
            'Дата оплаты Июль (факт)': service.date_july,
            'Сумма оплаты Июль (факт)': service.sum_july,
            'Дата оплаты Август (факт)': service.date_august,
            'Сумма оплаты Август (факт)': service.sum_august,
            'Дата оплаты Сентябрь (факт)': service.date_september,
            'Сумма оплаты Сентябрь (факт)': service.sum_september,
            'Дата оплаты Октябрь (факт)': service.date_october,
            'Сумма оплаты Октябрь (факт)': service.sum_october,
            'Дата оплаты Ноябрь (факт)': service.date_november,
            'Сумма оплаты Ноябрь (факт)': service.sum_november,
            'Дата оплаты Декабрь (факт)': service.date_december,
            'Сумма оплаты Декабрь (факт)': service.sum_december,
            'Дата оплаты Январь (факт) 1': service.date_january_two,
            'Сумма оплаты Январь (факт) 1': service.sum_january_two,
            '% исполнения (формула)': service.execution,
            'Остаток по контракту (формула)': service.contract_balance,
            'Color': getattr(service, 'color', '')
        } for service in services])

        # Приводим колонки к числовому типу данных
        df['НМЦК'] = pd.to_numeric(df['НМЦК'], errors='coerce')
        df['Цена контракта'] = pd.to_numeric(df['Цена контракта'], errors='coerce')
        df['Исполнение контракта (факт) (формула)'] = pd.to_numeric(df['Исполнение контракта (факт) (формула)'], errors='coerce')

        # Расчет итогов
        total_cost = df['НМЦК'].sum()
        total_certificate = df['Цена контракта'].sum()
        total_certificate_no = df['Исполнение контракта (факт) (формула)'].sum()

        keys_to_DataFrame = ('№ п/п', 'Наименование закупки', 'Статус',
                'Способ закупки', 'Инициатор закупки (ИТ/АХО)',
                'КЦСР', 'КОСГУ', 'ДопФК', 'НМЦК', 'Экономия',
                'Контрагент', 'Реестровый номер контракта (ЕИС)',
                'Номер контракта', 'Дата контракта', 'Окончание даты исполнения',
                'Цена контракта', 'Исполнение контракта (план) (формула)', 'Остаток контракта с предыдущего года',
                'Январь (план)', 'Февраль (план)', 'Март (план)', 'Апрель (план)',
                'Май (план)', 'Июнь (план)', 'Июль (план)', 'Август (план)', 'Сентябрь (план)',
                'Октябрь (план)', 'Ноябрь (план)', 'Декабрь (план)', 'Остаток контракта на следующий год',
                'Исполнение контракта (факт) (формула)', 'Оплачено в предыдущем году', 'Дата оплаты Январь (факт)',
                'Сумма оплаты Январь (факт)', 'Дата оплаты Февраль (факт)', 'Сумма оплаты Февраль (факт)',
                'Дата оплаты Март (факт)', 'Сумма оплаты Март (факт)', 'Дата оплаты Апрель (факт)',
                'Сумма оплаты Апрель (факт)', 'Дата оплаты Май (факт)', 'Сумма оплаты Май (факт)',
                'Дата оплаты Июнь (факт)', 'Сумма оплаты Июнь (факт)', 'Дата оплаты Июль (факт)',
                'Сумма оплаты Июль (факт)', 'Дата оплаты Август (факт)', 'Сумма оплаты Август (факт)',
                'Дата оплаты Сентябрь (факт)', 'Сумма оплаты Сентябрь (факт)', 'Дата оплаты Октябрь (факт)',
                'Сумма оплаты Октябрь (факт)', 'Дата оплаты Ноябрь (факт)', 'Сумма оплаты Ноябрь (факт)',
                'Дата оплаты Декабрь (факт)', 'Сумма оплаты Декабрь (факт)', 'Дата оплаты Январь (факт) 1',
                'Сумма оплаты Январь (факт) 1', '% исполнения (формула)', 'Остаток по контракту (формула)',
                'Color')

        value_to_DataFrame = {}
        for key in keys_to_DataFrame:
            if key == 'НМЦК':
                value_to_DataFrame[key] = total_cost
            elif key == 'Цена контракта':
                value_to_DataFrame[key] = total_certificate
            elif key == 'Исполнение контракта (факт) (формула)':
                value_to_DataFrame[key] = total_certificate_no
            else:
                value_to_DataFrame[key] = ''

        # Создание строки с итогами
        totals_row = pd.DataFrame([value_to_DataFrame])

        # # Создание строки с итогами
        # totals_row = pd.DataFrame([{
        #     '№ п/п': '',
        #     'Наименование закупки': '',
        #     'Статус': '',
        #     'Способ закупки': '',
        #     'Инициатор закупки (ИТ/АХО)': '',
        #     'КЦСР': '',
        #     'КОСГУ': '',
        #     'ДопФК': '',
        #     'НМЦК': total_cost,
        #     'Экономия': '',
        #     'Контрагент': '',
        #     'Реестровый номер контракта (ЕИС)': '',
        #     'Номер контракта': '',
        #     'Дата контракта': '',
        #     'Окончание даты исполнения': '',
        #     'Цена контракта': total_certificate,
        #     'Исполнение контракта (план) (формула)': '',
        #     'Остаток контракта с предыдущего года': '',
        #     'Январь (план)': '',
        #     'Февраль (план)': '',
        #     'Март (план)': '',
        #     'Апрель (план)': '',
        #     'Май (план)': '',
        #     'Июнь (план)': '',
        #     'Июль (план)': '',
        #     'Август (план)': '',
        #     'Сентябрь (план)': '',
        #     'Октябрь (план)': '',
        #     'Ноябрь (план)': '',
        #     'Декабрь (план)': '',
        #     'Остаток контракта на следующий год': '',
        #     'Исполнение контракта (факт) (формула)': total_certificate_no,
        #     'Оплачено в предыдущем году': '',
        #     'Дата оплаты Январь (факт)': '',
        #     'Сумма оплаты Январь (факт)': '',
        #     'Дата оплаты Февраль (факт)': '',
        #     'Сумма оплаты Февраль (факт)': '',
        #     'Дата оплаты Март (факт)': '',
        #     'Сумма оплаты Март (факт)': '',
        #     'Дата оплаты Апрель (факт)': '',
        #     'Сумма оплаты Апрель (факт)': '',
        #     'Дата оплаты Май (факт)': '',
        #     'Сумма оплаты Май (факт)': '',
        #     'Дата оплаты Июнь (факт)': '',
        #     'Сумма оплаты Июнь (факт)': '',
        #     'Дата оплаты Июль (факт)': '',
        #     'Сумма оплаты Июль (факт)': '',
        #     'Дата оплаты Август (факт)': '',
        #     'Сумма оплаты Август (факт)': '',
        #     'Дата оплаты Сентябрь (факт)': '',
        #     'Сумма оплаты Сентябрь (факт)': '',
        #     'Дата оплаты Октябрь (факт)': '',
        #     'Сумма оплаты Октябрь (факт)': '',
        #     'Дата оплаты Ноябрь (факт)': '',
        #     'Сумма оплаты Ноябрь (факт)': '',
        #     'Дата оплаты Декабрь (факт)': '',
        #     'Сумма оплаты Декабрь (факт)': '',
        #     'Дата оплаты Январь (факт) 1': '',
        #     'Сумма оплаты Январь (факт) 1': '',
        #     '% исполнения (формула)': '',
        #     'Остаток по контракту (формула)': '',
        #     'Color': ''
        # }])

        empty_rows_1 = pd.DataFrame([{
            '№ п/п': '№ п/п',
            'Наименование закупки': 'Наименование закупки',
            'Статус': 'Статус',
            'Способ закупки': 'Способ закупки',
            'Инициатор закупки (ИТ/АХО)': 'Инициатор закупки (ИТ/АХО)',
            'КЦСР': 'КЦСР',
            'КОСГУ': 'КОСГУ',
            'ДопФК': 'ДопФК',
            'НМЦК': 'НМЦК',
            'Экономия': 'Экономия',
            'Контрагент': 'Контрагент',
            'Реестровый номер контракта (ЕИС)': 'Реестровый номер контракта (ЕИС)',
            'Номер контракта': 'Номер контракта',
            'Дата контракта': 'Дата контракта',
            'Окончание даты исполнения': 'Окончание даты исполнения',
            'Цена контракта': 'Цена контракта',
            'Исполнение контракта (план) (формула)': 'Исполнение контракта (план) (формула)',
            'Остаток контракта с предыдущего года': 'Остаток контракта с предыдущего года',
            'Январь (план)': 'Январь (план)',
            'Февраль (план)': 'Февраль (план)',
            'Март (план)': 'Март (план)',
            'Апрель (план)': 'Апрель (план)',
            'Май (план)': 'Май (план)',
            'Июнь (план)': 'Июнь (план)',
            'Июль (план)': 'Июль (план)',
            'Август (план)': 'Август (план)',
            'Сентябрь (план)': 'Сентябрь (план)',
            'Октябрь (план)': 'Октябрь (план)',
            'Ноябрь (план)': 'Ноябрь (план)',
            'Декабрь (план)': 'Декабрь (план)',
            'Остаток контракта на следующий год': 'Остаток контракта на следующий год',
            'Исполнение контракта (факт) (формула)': 'Исполнение контракта (факт) (формула)',
            'Оплачено в предыдущем году': 'Оплачено в предыдущем году',
            'Дата оплаты Январь (факт)':'Январь (факт)',
            'Сумма оплаты Январь (факт)': 'Сумма оплаты Январь (факт)',
            'Дата оплаты Февраль (факт)': 'Февраль (факт)',
            'Сумма оплаты Февраль (факт)': 'Сумма оплаты Февраль (факт)',
            'Дата оплаты Март (факт)': 'Март (факт)',
            'Сумма оплаты Март (факт)': 'Сумма оплаты Март (факт)',
            'Дата оплаты Апрель (факт)': 'Апрель (факт)',
            'Сумма оплаты Апрель (факт)': 'Сумма оплаты Апрель (факт)',
            'Дата оплаты Май (факт)': 'Май (факт)',
            'Сумма оплаты Май (факт)': 'Сумма оплаты Май (факт)',
            'Дата оплаты Июнь (факт)': 'юнь (факт)',
            'Сумма оплаты Июнь (факт)': 'Сумма оплаты Июнь (факт)',
            'Дата оплаты Июль (факт)': 'Июль (факт)',
            'Сумма оплаты Июль (факт)': 'Сумма оплаты Июль (факт)',
            'Дата оплаты Август (факт)': 'Август (факт)',
            'Сумма оплаты Август (факт)': 'Сумма оплаты Август (факт)',
            'Дата оплаты Сентябрь (факт)': 'Сентябрь (факт)',
            'Сумма оплаты Сентябрь (факт)': 'Сумма оплаты Сентябрь (факт)',
            'Дата оплаты Октябрь (факт)': 'Октябрь (факт)',
            'Сумма оплаты Октябрь (факт)': 'Сумма оплаты Октябрь (факт)',
            'Дата оплаты Ноябрь (факт)': 'Ноябрь (факт)',
            'Сумма оплаты Ноябрь (факт)': 'Сумма оплаты Ноябрь (факт)',
            'Дата оплаты Декабрь (факт)': 'Декабрь (факт)',
            'Сумма оплаты Декабрь (факт)': 'Сумма оплаты Декабрь (факт)',
            'Дата оплаты Январь (факт) 1': 'Январь (факт) 1',
            'Сумма оплаты Январь (факт) 1': 'Сумма оплаты Январь (факт) 1',
            '% исполнения (формула)': '% исполнения (формула)',
            'Остаток по контракту (формула)': 'Остаток по контракту (формула)',
            'Color': 'Color'
        }])

        empty_rows_2 = pd.DataFrame([{
            '№ п/п': '',
            'Наименование закупки': '',
            'Статус': '',
            'Способ закупки': '',
            'Инициатор закупки (ИТ/АХО)': '',
            'КЦСР': '',
            'КОСГУ': '',
            'ДопФК': '',
            'НМЦК': '',
            'Экономия': '',
            'Контрагент': '',
            'Реестровый номер контракта (ЕИС)': '',
            'Номер контракта': '',
            'Дата контракта': '',
            'Окончание даты исполнения': '',
            'Цена контракта': '',
            'Исполнение контракта (план) (формула)': '',
            'Оплачено в предыдущем году': '',
            'Январь (план)': '',
            'Февраль (план)': '',
            'Март (план)': '',
            'Апрель (план)': '',
            'Май (план)': '',
            'Июнь (план)': '',
            'Июль (план)': '',
            'Август (план)': '',
            'Сентябрь (план)': '',
            'Октябрь (план)': '',
            'Ноябрь (план)': '',
            'Декабрь (план)': '',
            'Остаток контракта на следующий год': '',
            'Исполнение контракта (факт) (формула)': '',
            'Дата оплаты Январь (факт)':'Дата оплаты Январь (факт)',
            'Сумма оплаты Январь (факт)': 'Сумма оплаты Январь (факт)',
            'Дата оплаты Февраль (факт)': 'Дата оплаты Февраль (факт)',
            'Сумма оплаты Февраль (факт)': 'Сумма оплаты Февраль (факт)',
            'Дата оплаты Март (факт)': 'Дата оплаты Март (факт)',
            'Сумма оплаты Март (факт)': 'Сумма оплаты Март (факт)',
            'Дата оплаты Апрель (факт)': 'Дата оплаты Апрель (факт)',
            'Сумма оплаты Апрель (факт)': 'Сумма оплаты Апрель (факт)',
            'Дата оплаты Май (факт)': 'Дата оплаты Май (факт)',
            'Сумма оплаты Май (факт)': 'Сумма оплаты Май (факт)',
            'Дата оплаты Июнь (факт)': 'Дата оплаты Июнь (факт)',
            'Сумма оплаты Июнь (факт)': 'Сумма оплаты Июнь (факт)',
            'Дата оплаты Июль (факт)': 'Дата оплаты Июль (факт)',
            'Сумма оплаты Июль (факт)': 'Сумма оплаты Июль (факт)',
            'Дата оплаты Август (факт)': 'Дата оплаты Август (факт)',
            'Сумма оплаты Август (факт)': 'Сумма оплаты Август (факт)',
            'Дата оплаты Сентябрь (факт)': 'Дата оплаты Сентябрь (факт)',
            'Сумма оплаты Сентябрь (факт)': 'Сумма оплаты Сентябрь (факт)',
            'Дата оплаты Октябрь (факт)': 'Дата оплаты Октябрь (факт)',
            'Сумма оплаты Октябрь (факт)': 'Сумма оплаты Октябрь (факт)',
            'Дата оплаты Ноябрь (факт)': 'Дата оплаты Ноябрь (факт)',
            'Сумма оплаты Ноябрь (факт)': 'Сумма оплаты Ноябрь (факт)',
            'Дата оплаты Декабрь (факт)': 'Дата оплаты Декабрь (факт)',
            'Сумма оплаты Декабрь (факт)': 'Сумма оплаты Декабрь (факт)',
            'Дата оплаты Январь (факт) 1': 'Дата оплаты Январь (факт) 1',
            'Сумма оплаты Январь (факт) 1': 'Сумма оплаты Январь (факт) 1',
            '% исполнения (формула)': '',
            'Остаток по контракту (формула)': '',
            'Color': 'Color'
        }])

        # print('-----------------------')
        # print(f"Первое {empty_rows_1.columns}")
        # print(f"Первое {empty_rows_2.columns}")
        # print(f"Второе {df.columns}")
        # print(f"Третье {totals_row.columns}")

        # exit()

        # # Перемещаем все DataFrame в одну структуру с одинаковыми столбцами
        # empty_rows_1 = empty_rows_1[columns]
        # empty_rows_2 = empty_rows_2[columns]
        # df = df[columns]
        # totals_row = totals_row[columns]

        # Добавление строки с итогами в DataFrame
        df = pd.concat([empty_rows_1, empty_rows_2, df, totals_row], ignore_index=True)
        # df = pd.concat([df, totals_row], ignore_index=True)

        # Создаем файл Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Services', startcol=0)

            # Настройка стиля
            worksheet = writer.sheets['Services']

            # Определяем стиль для границ
            border_style = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

            # Определяем стиль для заливки желтым цветом
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            import re
            def is_valid_hex_color(color):
                # Регулярное выражение для проверки формата цвета
                pattern = r'^#[0-9A-Fa-f]{6}$'
                return bool(re.match(pattern, color))

            # Применяем цвет к ячейкам, где он задан
            for row_num in range(2, worksheet.max_row):  # Пропускаем заголовки
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)

                    color = df.iloc[row_num]['Color']  # Сопоставление индексов DataFrame
                    # Применяем цвет только к ячейкам данных
                    if color and is_valid_hex_color(color):
                        cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")

            # Удаляем столбец Color из Excel-файла
            worksheet.delete_cols(df.columns.get_loc("Color") + 1)

            # Получаем актуальные размеры после удаления столбца
            max_row = worksheet.max_row
            max_column = worksheet.max_column

            # Применяем границы ко всем ячейкам
            for row_num in range(1, max_row):  # Пропускаем заголовки
                for col_num in range(1, max_column + 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.border = border_style

            # # Применяем границы к заголовкам
            # for col_num in range(1, worksheet.max_column + 1):
            #     cell = worksheet.cell(row=1, column=col_num)
            #     cell.border = border_style

            # Применяем желтый цвет к строке с итогами
            totals_row_num = max_row
            for col_num in range(1, max_column + 1):
                cell = worksheet.cell(row=totals_row_num, column=col_num)
                cell.fill = yellow_fill
                cell.border = border_style

        output.seek(0)

        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        import datetime
        date = str(datetime.datetime.now().date())

        # Получаем текущую директорию проекта
        project_dir = os.path.dirname(os.path.abspath(__file__))

        # Строим путь к папке file внутри проекта
        file_dir = os.path.join(project_dir, 'file')

        # Создаем директорию file, если она не существует
        os.makedirs(file_dir, exist_ok=True)

        # Строим полный путь к файлу
        file_path = os.path.join(file_dir, f'services_{sid}_{date}.xlsx')

        # Сохраняем файл на диск
        with open(file_path, 'wb') as f:
            f.write(output.read())

        # print('POPAL FILE', project_dir)
        # exit()

        # Открытие файла с openpyxl для объединения ячеек
        wb = load_workbook(file_path)
        ws = wb.active

        # Объединение ячеек
        for col in range(1, 62):  # 1 - это 'A', 2 - это 'B' и т.д.
            if 34 <= col <= 59 and col % 2 == 0: # Исключаем 33, 35, 37 и т.д.
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            elif not (34 <= col <= 59):
                ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

        # Выравниваем текст по центру
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=62):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Сохранение изменений
        wb.save(file_path)

        # Возвращаем путь к файлу для отправки клиенту
        filename = f"services_{sid}_{date}.xlsx"
        file_url = f"/file/{filename}"

        # # Emit the success event with file URL
        # socketio.emit('export_success', {'file_url': file_url, 'filename': filename}, room=sid)

        return file_url, filename

        # # Преобразуем файл в base64, чтобы отправить его через WebSocket
        # file_data = base64.b64encode(output.read()).decode('utf-8')
        # # return file_data

        # print(f"Отправка данных клиенту: {file_data}, имя файла: services.xlsx")

        # return file_data

    except Exception as e:
        print(f"Ошибка в задаче для сессии {sid}: {e}")
        # Работа с channel_layer через sync_to_async
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "export_error",
            {
                "type": "export.error",
                "message": str(e),
            }
        )
